"""
Excel文档操作模块 - AI可调用的Excel文档编辑工具
使用openpyxl库实现.xlsx/.xls文件的读取/编辑
"""
from .基类 import 操作结果, 操作基类


class 替换Excel文本(操作基类):
    名称 = "替换Excel文本"
    描述 = "在.xlsx/.xls文件中查找并替换单元格文本，支持全部替换或仅替换第一个"
    参数结构 = {
        "路径": {"类型": "字符串", "必填": True, "说明": "Excel文档(.xlsx/.xls)路径"},
        "旧文本": {"类型": "字符串", "必填": True, "说明": "要查找的文本"},
        "新文本": {"类型": "字符串", "必填": True, "说明": "替换为的文本"},
        "全部替换": {"类型": "布尔", "必填": False, "说明": "True=替换全部匹配，False=仅替换第一个，默认True"}
    }

    def 执行(self, 参数: dict) -> 操作结果:
        路径 = 参数.get("路径", "")
        旧文本 = 参数.get("旧文本", "")
        新文本 = 参数.get("新文本", "")
        全部 = 参数.get("全部替换", True)
        if not 路径 or not 旧文本:
            return 操作结果.失败("路径和旧文本不能为空")
        try:
            from openpyxl import load_workbook
            wb = load_workbook(路径)
            替换数 = 0
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None and 旧文本 in str(cell.value):
                            cell.value = str(cell.value).replace(旧文本, 新文本)
                            替换数 += 1
                            if not 全部:
                                wb.save(路径)
                                return 操作结果.成功(f"已替换1处", {"操作类型": "替换Excel文本", "替换数": 1})
            if 替换数 > 0:
                wb.save(路径)
                return 操作结果.成功(f"已替换{替换数}处", {"操作类型": "替换Excel文本", "替换数": 替换数})
            return 操作结果.失败(f"未找到文本: {旧文本}")
        except Exception as e:
            return 操作结果.失败(f"替换Excel文本失败: {e}")
